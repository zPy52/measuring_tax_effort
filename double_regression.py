from __future__ import annotations
from dataclasses import dataclass
from typing import List, Tuple
from math import sqrt
from csv import DictReader

from openpyxl import load_workbook, Workbook
import numpy as np
from sklearn.linear_model import LinearRegression

# List of countries to study.
exemplary_countries = ('Singapore', 'Ireland', 'Korea, Rep.')
example_countries = ('Spain', 'France', 'Italy', 'United Kingdom', 'United States', 'Canada', 
                     'Netherlands', 'Belgium', 'Portugal', 'Greece', 'Japan')

# Helper functions to use later.
PHI = (1 + sqrt(5)) / 2
def tax_effort(tax_burden: float, unemployment: float, gdp_ppp_per_capita: int | float):
    for percentage, parameter_title in {(tax_burden, 'tax burden'), (unemployment, 'unemployment')}:
        if percentage > 1 or percentage < 0:
            raise ValueError(f'Invalid {parameter_title}: it must be a (positive) percentage represented as a number from 0 to 1, both included.')
    
    return tax_burden / ((1 - tax_burden) * (1 - unemployment) * (gdp_ppp_per_capita ** PHI))

# Helper dataclass for countries' general data.
@dataclass
class Country:
    country_name: str
    year: int

    tax_burden: float
    unemployment: float
    gdp_ppp_per_capita: float

    real_gdp_per_capita: float

    @property
    def tax_effort(self) -> float:
        return tax_effort(self.tax_burden, self.unemployment, self.gdp_ppp_per_capita)

# Helper dataclass for countries' real GDP per capita data.
@dataclass
class RealGDP:
    __slots__ = 'country_name', 'year', 'real_gdp'

    country_name: str
    year: int

    real_gdp: float
    

def get_median(series: List[int | float]) -> float:
    series.sort()
    length = len(series)
    if length % 2 == 0:
        return (series[length // 2] + series[length // 2 - 1]) / 2
    else:
        return series[length // 2]

def normalize(series: List[int | float] | Tuple[int | float], element: int | float = None) -> List[float] | float:
    if not isinstance(series, (list, tuple)):
        raise ValueError()
    
    mean = sum(series) / len(series)

    std = sqrt(sum([(data_point - mean) ** 2 for data_point in series]) / len(series))

    return (element - mean) / std if element is not None else [(data_point - mean) / std for data_point in series]


def get_common_min(series1: List[Country] | Tuple[Country], series2: List[Country] | Tuple[Country]):
    initially_upper = series1 if series1[0].real_gdp_per_capita > series2[0].real_gdp_per_capita else series2
    initially_lower = series1 if series1[0].real_gdp_per_capita <= series2[0].real_gdp_per_capita else series2

    AMIN = min(initially_upper, key=lambda country: (country.real_gdp_per_capita, country.year))
    BMIN = min(initially_lower, key=lambda country: (abs(country.real_gdp_per_capita - AMIN.real_gdp_per_capita), country.year))

    return {AMIN.country_name: AMIN, BMIN.country_name: BMIN}

def get_common_max(series1: List[Country] | Tuple[Country], series2: List[Country] | Tuple[Country]):
    finally_upper = series1 if series1[-1].real_gdp_per_capita > series2[-1].real_gdp_per_capita else series2
    finally_lower = series1 if series1[-1].real_gdp_per_capita <= series2[-1].real_gdp_per_capita else series2

    AMAX = max(finally_lower, key=lambda country: (country.real_gdp_per_capita, country.year))
    BMAX = min(finally_upper, key=lambda country: (abs(country.real_gdp_per_capita - AMAX.real_gdp_per_capita), country.year))

    return {AMAX.country_name: AMAX, BMAX.country_name: BMAX}


### LOADING DATA ###
wb = load_workbook('datasets/World_Bank/WB_Unemployment.xlsx')

ws = wb.active

countries = {}

### Unemployment rates ###
for row in range(1, ws.max_row + 1):
    country_name = ws.cell(row=row, column=1).value

    last_year = None
    column = 4
    while last_year != 2021:
        column += 1
        
        last_year = int(ws.cell(row=1, column=column).value)

        countries.setdefault(last_year, {})
        countries[last_year].setdefault(country_name, {'tax_burden': None, 'unemployment': None, 
                                                        'gdp_ppp_per_capita': None, 'real_gdp_per_capita': None})

        value = ws.cell(row=row, column=column).value
        try:
            countries[last_year][country_name]['unemployment'] = float(value) / 100
        except Exception:
            continue

wb.close()

### Tax burdens ###
with open('datasets/Our_World_In_Data/OWID_Total_Tax_Revenues_GDP.csv', 'r') as file:
    data = DictReader(file)

    for row in data:
        country_name = row['Entity']
        country_name = country_name.replace('South Korea', 'Korea, Rep.')

        last_year = int(row['Year'])

        countries.setdefault(last_year, {})
        countries[last_year].setdefault(country_name, {'tax_burden': None, 'unemployment': None, 
                                                        'gdp_ppp_per_capita': None, 'real_gdp_per_capita': None})

        countries[last_year][country_name]['tax_burden'] = float(row[r'Total tax revenue (% of GDP) (ICTD (2021))']) / 100

### GDPs Per Capita ###
wb = load_workbook('datasets/International_Monetary_Fund/IMF_GDP_Per_Capita_PPP.xlsx')

ws = wb.active

for row in range(1, ws.max_row + 1):
    country_name = ws.cell(row=row, column=1).value
    
    if isinstance(country_name, str):
        country_name = country_name.replace('Korea, Republic of', 'Korea, Rep.')

    last_year = None
    column = 1
    while last_year != 2021:
        column += 1

        last_year = int(ws.cell(row=1, column=column).value)
        
        countries.setdefault(last_year, {})
        countries[last_year].setdefault(country_name, {'tax_burden': None, 'unemployment': None, 
                                                        'gdp_ppp_per_capita': None, 'real_gdp_per_capita': None})

        value = ws.cell(row=row, column=column).value
        try:
            countries[last_year][country_name]['gdp_ppp_per_capita'] = float(value)
        except Exception:
            continue

wb.close()

### Real GDPs Per Capita ###
with open('datasets/Our_World_In_Data/OWID_GDP_Per_Capita_In_US_Dollar_World_Bank.csv', 'r') as file:
    data = DictReader(file)

    for row in data:
        country_name = row['Entity']
        country_name = country_name.replace('South Korea', 'Korea, Rep.')

        last_year = int(row['Year'])

        countries.setdefault(last_year, {})
        countries[last_year].setdefault(country_name, {'tax_burden': None, 'unemployment': None, 
                                                        'gdp_ppp_per_capita': None, 'real_gdp_per_capita': None})

        countries[last_year][country_name]['real_gdp_per_capita'] = float(row['GDP per capita (constant 2015 US$)'])

data = {}
real_gdp_history = {}
for year in countries:
    for country_name in countries[year]:
        data.setdefault(country_name, [])
        real_gdp_history.setdefault(country_name, [])

        if countries[year][country_name]['real_gdp_per_capita'] is not None:
            real_gdp_history[country_name].append(RealGDP(
                country_name=country_name, 
                year=year, 
                real_gdp=countries[year][country_name]['real_gdp_per_capita']    
            ))

        if (countries[year][country_name]['tax_burden'] is None
            or countries[year][country_name]['unemployment'] is None
            or countries[year][country_name]['gdp_ppp_per_capita'] is None
            or countries[year][country_name]['real_gdp_per_capita'] is None):
            continue

        data[country_name].append(Country(
            country_name=country_name, 
            year=year,
            tax_burden=countries[year][country_name]['tax_burden'],
            unemployment=countries[year][country_name]['unemployment'],
            gdp_ppp_per_capita=countries[year][country_name]['gdp_ppp_per_capita'],
            real_gdp_per_capita=countries[year][country_name]['real_gdp_per_capita'],
        ))

### REMOVING MISSING DATA ###
for country_name in set(data.keys()):
    if len(data[country_name]):
        data[country_name].sort(key=lambda country: country.year)
    else:
        del data[country_name]

    if len(real_gdp_history[country_name]):
        real_gdp_history[country_name].sort(key=lambda country: country.year)
    else:
        del real_gdp_history[country_name]

### USING THE DATA ###

# Calculating lambdas.
def core(example_country):
    lambdas = {}
    for exemplary_country in exemplary_countries:
        common_min = get_common_min(data[exemplary_country], data[example_country])
        common_max = get_common_max(data[exemplary_country], data[example_country])

        example_country_difference = common_max[example_country].year - common_min[example_country].year
        exemplary_country_difference = common_max[exemplary_country].year - common_min[exemplary_country].year

        try:
            lambdas[exemplary_country] = (example_country_difference / exemplary_country_difference, exemplary_country_difference)
        except ZeroDivisionError:
            lambdas[exemplary_country] = (0, 0)

    # Calculating weights.
    # Median tax efforts, median real GDP per capita and the inverse of lambda intervals.
    median_tax_efforts = {
        country_name:
        normalize([instance.tax_effort for instance in data[country_name]], 
                get_median([instance.tax_effort for instance in data[country_name]])) 
        
        for country_name in exemplary_countries
    }

    tax_burden_goal = sum([
        
        (get_median([instance.tax_burden for instance in data[exemplary_country]]) 
            * (1 - get_median([instance.unemployment for instance in data[example_country]])) 
            * (get_median([instance.gdp_ppp_per_capita for instance in data[example_country]]) ** PHI))
        /
        ((1 - get_median([instance.tax_burden for instance in data[exemplary_country]])) 
            * (1 - get_median([instance.unemployment for instance in data[exemplary_country]]))
            * ((get_median([instance.gdp_ppp_per_capita for instance in data[exemplary_country]])) ** PHI)
        +
            get_median([instance.tax_burden for instance in data[exemplary_country]])
            * (1 - get_median([instance.unemployment for instance in data[example_country]]))
            * (get_median([instance.gdp_ppp_per_capita for instance in data[example_country]]) ** PHI))
        
        for exemplary_country in exemplary_countries]) / len(exemplary_countries)

    resulting_median_tax_effort = tax_effort(tax_burden_goal, 
                                      get_median([instance.unemployment for instance in data[example_country]]), 
                                      get_median([instance.gdp_ppp_per_capita for instance in data[example_country]]))

    median_tax_efforts[example_country] = normalize([instance.tax_effort for instance in data[example_country]], 
                resulting_median_tax_effort) 

    try:
        interval_inverses = [1 / lambdas[country_name][1] for country_name in exemplary_countries]
    except ZeroDivisionError:
        return None

    intervals = sorted([
        (
        country_name,
        normalize(interval_inverses, 
                1 / lambdas[country_name][1])
        )
        
        for country_name in exemplary_countries
    ], key=lambda x: x[1])

    intervals = [(instance[0], instance[1] + abs(intervals[0][1])) for instance in intervals]

    distances = []
    for exemplary_country in exemplary_countries:
        full_distance = sqrt(
                (median_tax_efforts[exemplary_country] - median_tax_efforts[example_country]) ** 2 
                + dict(intervals)[exemplary_country] ** 2
            )

        distances.append((exemplary_country, full_distance))

    total_distance = sum([instance[1] for instance in distances])

    distances.sort(key=lambda instance: instance[1])

    raw_weights = (reversed([instance[1] / total_distance for instance in distances]))

    weights = dict(zip([instance[0] for instance in distances], raw_weights))

    FINAL_LAMBDA = sum([lambdas[exemplary_country][0] * weights[exemplary_country] for exemplary_country in exemplary_countries])

    x = np.array([
        instance.year - real_gdp_history[example_country][0].year for instance in real_gdp_history[example_country]
    ]).reshape(-1, 1)

    y = np.array([instance.real_gdp for instance in real_gdp_history[example_country]])

    approximate_alpha_function = LinearRegression()
    approximate_alpha_function.fit(x, y)

    alpha_star = lambda x: FINAL_LAMBDA * approximate_alpha_function.coef_[0] * x + approximate_alpha_function.intercept_

    return {'country_name': example_country, 
            'years': {country_name: lambdas[country_name][1] for country_name in exemplary_countries},
            'weights': {country_name: f'{weights[country_name]:,.2%}' for country_name in weights},
            'actual_real_gdp_per_capita': f'${real_gdp_history[example_country][-1].real_gdp:,.2f}', 
            'estimation': f'${alpha_star(len(real_gdp_history[example_country]) - 1):,.2f}',
            'tax_burden_relation': f'{tax_burden_goal / get_median([instance.tax_burden for instance in data[example_country]]):,.2%}'}

wb = Workbook()

ws = wb.active

for index, example_country in enumerate(example_countries):
    core_data = core(example_country)
    ws.cell(row=index + 3, column=1).value = example_country
    
    if core_data is None:
        ws.cell(row=index + 3, column=2).value = '---'
        continue

    ws.cell(row=index + 3, column=2).value = core_data['estimation']
    ws.cell(row=index + 3, column=3).value = core_data['actual_real_gdp_per_capita']
    ws.cell(row=index + 3, column=4).value = core_data['tax_burden_relation']

    for i, exemplary_country in zip(range(0, len(core_data['years']) * 2, 2), core_data['years'].keys()):
        ws.cell(row=1, column=5 + i).value = exemplary_country
        ws.cell(row=1, column=6 + i).value = exemplary_country
        
        ws.cell(row=index + 3, column=5 + i).value = core_data['weights'][exemplary_country]
        ws.cell(row=index + 3, column=6 + i).value = core_data['years'][exemplary_country]


#wb.save('table_y.2.1.xlsx')