# Import proper libraries.
from typing import List, Tuple
from math import sqrt
from openpyxl import load_workbook
from csv import DictReader

import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
import matplotlib.pyplot as plt

# Define a function to find a certain element given.
def find_element_by_year(series: List[Tuple[int, float]], year: int) -> Tuple[int, float]:
    for element in series:
        if element[0] == year:
            return element[1]
    else:
        return None

# Dictionary to hold every country's data.
countries = {}

# Series of countries to analyse.
countries_to_analyse = ('Singapore', 'Ireland', 'Luxembourg', 'Canada', 'Switzerland', 'Korea, Rep.', 
                        'Spain', 'France', 'Italy', 'United Kingdom', 'United States', 'Netherlands', 
                        'Belgium', 'Portugal', 'Greece', 'Japan')


# These are the low-effort, rich nations and high-effort, poor nations.
"""countries_to_analyse = ('Singapore', 'Ireland', 'Luxembourg', 'Switzerland', 'Korea, Rep.',
                            'Bangladesh', 'South Africa', 'Brazil', 'India', 'Cambodia')"""


# Set a default's dictionary structure inside 'countries' to save pertinent data.
for country_name in countries_to_analyse:
    countries[country_name] = {'tax_burden': [], 'unemployment': [], 'gdp_ppp': []}

### Tax burdens ###
# Load tax burden data into the 'countries' dictionary.
with open('datasets/Our_World_In_Data/OWID_Total_Tax_Revenues_GDP.csv', 'r') as file:
    data = DictReader(file)

    for row in data:
        country_name = row['Entity']
        country_name = country_name.replace('South Korea', 'Korea, Rep.')

        if country_name in countries_to_analyse:
            countries[country_name]['tax_burden'].append((
                int(row['Year']),
                float(row[r'Total tax revenue (% of GDP) (ICTD (2021))']) / 100
            ))

### Unemployment rates ###
# Load unemployment rates data into the 'countries' dictionary.
wb = load_workbook('datasets/World_Bank/WB_Unemployment.xlsx')

ws = wb.active

for row in range(1, ws.max_row + 1):
    country_name = ws.cell(row=row, column=1).value
    
    if country_name not in countries_to_analyse:
        continue

    last_year = None
    column = 4
    while last_year != 2021:
        column += 1
        
        last_year = int(ws.cell(row=1, column=column).value)

        value = ws.cell(row=row, column=column).value
        try:
            countries[country_name]['unemployment'].append((
                last_year,
                float(value) / 100
            ))
        except Exception:
            continue

wb.close()

### GDPs Per Capita (PPP) ###
# Load GDPs per capita (PPP) data into the 'countries' dictionary.
wb = load_workbook('datasets/International_Monetary_Fund/IMF_GDP_Per_Capita_PPP.xlsx')

ws = wb.active

for row in range(1, ws.max_row + 1):
    
    country_name = ws.cell(row=row, column=1).value
    
    if isinstance(country_name, str):
        country_name = country_name.replace('Korea, Republic of', 'Korea, Rep.')

    if country_name not in countries_to_analyse:
        continue

    last_year = None
    column = 1
    while last_year != 2021:
        column += 1

        last_year = int(ws.cell(row=1, column=column).value)

        value = ws.cell(row=row, column=column).value
        try:
            countries[country_name]['gdp_ppp'].append((last_year, float(value)))
        except Exception:
            continue

wb.close()

### Building Missing Values ###
#  Define a constant PHI with the golden number and a lambda function to calculate
# the tax effort with the formula used on the paper.
PHI = (1 + sqrt(5)) / 2 
tax_effort = lambda tax_burden, unemployment, gdp_ppp: tax_burden / ((1 - tax_burden) * (1 - unemployment) * (gdp_ppp ** PHI))


### PLOT ###
#  Years and its index (from 1950 to 2022, both included, the year 1950 would 
# have index 1 and 2022 would have index 73).
year_index = {year: i + 1 for i, year in enumerate(range(1950, 2022 + 1))}

#  There are available two modes: 'regression', 'average' & 'median'. The first
# one makes a linear regression if REG_DEGREE is 1 and a polynomial regression
# of degree REG_DEGREE otherwise. The second one calculates the average of
# tax effort values and draws a line with it through the whole chart. The third
# one just returns the median of the data array.
mode = 'median'

#  Degree of the linear/polynomial regression to make. Must be an integer
# greater than zero.
REG_DEGREE = 1

# Regression case.
if mode == 'regression':
    plt.title('Linear Regression of Tax Efforts' if REG_DEGREE == 1 
                    else f'Polynomial Regression of Degree {REG_DEGREE} of Tax Efforts')

    print(f'Average Regression (Degree {REG_DEGREE}) Results:')

# Average case.
elif mode == 'average':
    plt.title('Average of Tax Efforts')

    print('Average Value Results:')

# Median case.
elif mode == 'median':
    plt.title('Median of Tax Efforts')

    print('Median Values Results:')


# Load and plot data from every country listed in 'countries_to_analyse'.
for country_name in countries:
    x = []
    y = []

    for year in year_index:
        #  Look for every element (tuple of two values, the year and country's tax effort at the time) 
        # through the function defined earlier.
        tax_burden = find_element_by_year(countries[country_name]['tax_burden'], year)
        unemployment = find_element_by_year(countries[country_name]['unemployment'], year)
        gdp_ppp = find_element_by_year(countries[country_name]['gdp_ppp'], year)

        # If any of the needed variables are missing (is None), then skip and leave blank.
        if None in {tax_burden, unemployment, gdp_ppp}:
            continue

        #  Otherwise, append to the X-axis list of values the year's index and to the Y-axis
        # list of value the tax effort multiplied by a constant (to avoid numbers of the 
        # sort of 6.01e-07).
        x.append(year_index[year])
        y.append((10 ** 10) * tax_effort(tax_burden, unemployment, gdp_ppp))

    #  Plot dots: X-axis is year's index and Y-axis is its tax effort then.
    plt.scatter(x, y, label=country_name)

    # Regression case.
    if mode == 'regression':
        X = np.array(x).reshape(-1, 1)
        Y = np.array(y)
        
        X_poly = PolynomialFeatures(degree=REG_DEGREE).fit_transform(X)
        model = LinearRegression()
        model.fit(X_poly, Y)

        x_line = np.linspace(0, len(year_index), 10_000).reshape(-1, 1)
        y_line = model.predict(PolynomialFeatures(degree=REG_DEGREE).fit_transform(x_line))
        
        plt.plot(x_line, y_line, label=f'Regression of {country_name}')

        expected_tax_effort = model.predict(PolynomialFeatures(degree=REG_DEGREE)
                                    .fit_transform(np.arange(len(year_index)).reshape(-1, 1)))

        average_predicted_tax_effort = sum(expected_tax_effort) / len(expected_tax_effort)

        print(f'  - {country_name}: {average_predicted_tax_effort}')
    

    # Average case.
    if mode == 'average':
        average = sum(y) / len(y)
        
        x_line = [0, len(year_index)]
        y_line = [average] * 2

        plt.plot(x_line, y_line, label=f'Average of {country_name}')

        print(f'  - {country_name}: {round(average, 2)}')

    # Median case.
    if mode == 'median':
        ordered_datapoints = sorted(y)

        if len(ordered_datapoints) % 2 == 1:
            median = ordered_datapoints[len(ordered_datapoints) // 2]
        else:
            median = (ordered_datapoints[len(ordered_datapoints) // 2 - 1] + ordered_datapoints[len(ordered_datapoints) // 2]) / 2
        
        x_line = [0, len(year_index)]
        y_line = [median] * 2

        plt.plot(x_line, y_line, label=f'Median of {country_name}')

        print(f'  - {country_name}: {round(median, 2)}')

# Style plotted figure and show on screen.
plt.xlabel('Year (as index)')
plt.ylabel('Tax Effort ')

plt.legend()

plt.show()